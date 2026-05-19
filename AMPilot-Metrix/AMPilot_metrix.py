import argparse
import ast
import csv
import json
import math
import re
from pathlib import Path

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
TRAJECTORY_KEY = "Trajectory"

EXPECTED_HORIZON_STEPS = 4
LLM_DISTANCE_THRESHOLD = 1200.0

MODEL_SPECS = [
    ("MiningCoT7B", "MiningCoT7B_{scenario}_predict_1000.json", "MiningCoT_{scenario}_test_1000.json"),
    ("MiningCoT13B", "MiningCoT13B_{scenario}_predict_1000.json", "MiningCoT_{scenario}_test_1000.json"),
    (
        "MiningNoCoT7B",
        "MiningNoCoT7B_{scenario}_predict_1000.json",
        "MiningCoT_{scenario}_test_1000_NoCoT.json",
    ),
    (
        "MiningNoCoT13B",
        "MiningNoCoT13B_{scenario}_predict_1000.json",
        "MiningCoT_{scenario}_test_1000_NoCoT.json",
    ),
    ("LC-LLM_CoT7B", "LC-LLM_CoT7B_{scenario}_predict_1000.json", "LC-LLM_CoT_{scenario}_test_1000.json"),
    ("LC-LLM_CoT13B", "LC-LLM_CoT13B_{scenario}_predict_1000.json", "LC-LLM_CoT_{scenario}_test_1000.json"),
    (
        "LC-LLMNoCoT7B",
        "LC-LLMNoCoT7B_{scenario}_predict_1000.json",
        "LC-LLM_CoT_{scenario}_test_1000_NoCoT.json",
    ),
    (
        "LC-LLMNoCoT13B",
        "LC-LLMNoCoT13B_{scenario}_predict_1000.json",
        "LC-LLM_CoT_{scenario}_test_1000_NoCoT.json",
    ),
]

ALL_OUTPUT_MODEL_SPECS = [
    ("MiningCoT7B", "MiningCoT7B_{scenario}_predict_5000.json", "MiningCoT_{scenario}_test_5000.json"),
    ("MiningCoT13B", "MiningCoT13B_{scenario}_predict_5000.json", "MiningCoT_{scenario}_test_5000.json"),
    (
        "MiningNoCoT7B",
        "MiningNoCoT7B_{scenario}_predict_5000.json",
        "MiningCoT_{scenario}_test_5000_NoCoT.json",
    ),
    (
        "MiningNoCoT13B",
        "MiningNoCoT13B_{scenario}_predict_5000.json",
        "MiningCoT_{scenario}_test_5000_NoCoT.json",
    ),
    ("LC-LLM_CoT7B", "LC-LLM_CoT7B_{scenario}_predict_5000.json", "LC-LLM_CoT_{scenario}_test_5000.json"),
    ("LC-LLM_CoT13B", "LC-LLM_CoT13B_{scenario}_predict_5000.json", "LC-LLM_CoT_{scenario}_test_5000.json"),
    (
        "LC-LLMNoCoT7B",
        "LC-LLMNoCoT7B_{scenario}_predict_5000.json",
        "LC-LLM_CoT_{scenario}_test_5000_NoCoT.json",
    ),
    (
        "LC-LLMNoCoT13B",
        "LC-LLMNoCoT13B_{scenario}_predict_5000.json",
        "LC-LLM_CoT_{scenario}_test_5000_NoCoT.json",
    ),
]

SCENARIO0_MODEL_SPECS = [
    ("MiningCoT7B", "MiningCoT7B_predict_5000.json", "MiningCoT_test_5000.json"),
    ("MiningCoT13B", "MiningCoT13B_predict_5000.json", "MiningCoT_test_5000.json"),
    ("MiningNoCoT7B", "MiningNoCoT7B_predict_5000.json", "Mining_noCoT_test_5000.json"),
    (
        "MiningNoCoT13B",
        "MiningNoCoT13B_predict_5000.json",
        "Mining_noCoT_test_5000.json",
    ),
    ("LC-LLM_CoT7B", "LC-LLM_CoT7B_predict_5000.json", "LC-LLM_CoT_test_5000.json"),
    ("LC-LLM_CoT13B", "LC-LLM_CoT13B_predict_5000.json", "LC-LLM_CoT_test_5000.json"),
    ("LC-LLMNoCoT7B", "LC-LLMNoCoT7B_predict_5000.json", "LC-LLM_noCoT_test_5000.json"),
    ("LC-LLMNoCoT13B", "LC-LLMNoCoT13B_predict_5000.json", "LC-LLM_noCoT_test_5000.json"),
]

INTENTION_CLASSES = ["straight", "left", "right"]
LOAD_STATES = ["full_loaded", "not_full_loaded"]
NEAREST_MINING_TRUCK_FULLY_LOADED_PHRASE = "the nearest mining truck is fully loaded"
AMPILOT_METRIX_COLUMNS = [
    "scenario",
    "group",
    "model",
    "sample_count",
    "RMSE",
    "ADE",
    "AHE",
    "FDE",
    "FHE",
    "rmse_x_t1",
    "rmse_x_t2",
    "rmse_x_t3",
    "rmse_x_t4",
    "rmse_y_t1",
    "rmse_y_t2",
    "rmse_y_t3",
    "rmse_y_t4",
    "load_state_evaluation",
    "intention_prediction",
]


def load_json(path):
    with open(path, "r", encoding="utf-8") as file:
        return json.load(file)


def extract_balanced_brackets(text, start_index):
    if start_index < 0 or start_index >= len(text) or text[start_index] != "[":
        return None

    depth = 0
    for index in range(start_index, len(text)):
        char = text[index]
        if char == "[":
            depth += 1
        elif char == "]":
            depth -= 1
            if depth == 0:
                return text[start_index : index + 1]
    return None


def extract_trajectory_block_from_text(text, trajectory_key=TRAJECTORY_KEY):
    if not isinstance(text, str):
        return None

    patterns = [
        rf'"{trajectory_key}"\s*:\s*\[',
        rf"{trajectory_key}\s*:\s*\[",
        rf"-\s*{trajectory_key}\s*:\s*\[",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            bracket_start = text.find("[", match.start())
            block = extract_balanced_brackets(text, bracket_start)
            if block is not None:
                return block

    quoted_patterns = [
        rf'"?{trajectory_key}"?\s*:\s*[\'"]\s*\[',
        rf"-\s*{trajectory_key}\s*:\s*[\'\"]\s*\[",
        rf"Final Answer:\s*-?\s*{trajectory_key}\s*:\s*[\'\"]\s*\[",
    ]
    for pattern in quoted_patterns:
        match = re.search(pattern, text)
        if match:
            bracket_start = text.find("[", match.start())
            block = extract_balanced_brackets(text, bracket_start)
            if block is not None:
                return block
    return None


def normalize_point(point):
    if not isinstance(point, (list, tuple)) or len(point) < 2:
        return None

    x_value, y_value = point[0], point[1]
    if not isinstance(x_value, (int, float)) or not isinstance(y_value, (int, float)):
        return None
    if not (math.isfinite(float(x_value)) and math.isfinite(float(y_value))):
        return None
    return [float(x_value), float(y_value)]


def normalize_trajectory(trajectory):
    if not isinstance(trajectory, list) or not trajectory:
        return None

    normalized = []
    for point in trajectory:
        normalized_point = normalize_point(point)
        if normalized_point is None:
            return None
        normalized.append(normalized_point)
    return normalized


def parse_trajectory_from_output(output_content, trajectory_key=TRAJECTORY_KEY):
    if isinstance(output_content, dict):
        return normalize_trajectory(output_content.get(trajectory_key))

    if not isinstance(output_content, str):
        return None

    try:
        parsed_json = json.loads(output_content)
    except json.JSONDecodeError:
        parsed_json = None

    if isinstance(parsed_json, dict):
        trajectory = normalize_trajectory(parsed_json.get(trajectory_key))
        if trajectory is not None:
            return trajectory

    trajectory_block = extract_trajectory_block_from_text(output_content, trajectory_key=trajectory_key)
    if trajectory_block is None:
        return None

    try:
        return normalize_trajectory(ast.literal_eval(trajectory_block))
    except (SyntaxError, ValueError):
        return None


def angle_difference(predicted_angles, label_angles):
    return np.abs(np.arctan2(np.sin(predicted_angles - label_angles), np.cos(predicted_angles - label_angles)))


def compute_heading(trajectories):
    deltas = np.diff(trajectories, axis=1)
    return np.arctan2(deltas[:, :, 1], deltas[:, :, 0])


def evaluate_aligned(aligned):
    if not aligned:
        raise ValueError("No valid aligned trajectories are available.")

    label = np.asarray([record["label"] for record in aligned], dtype=float)
    prediction = np.asarray([record["prediction"] for record in aligned], dtype=float)
    diff = prediction - label
    point_error = np.linalg.norm(diff, axis=2)
    final_error = point_error[:, -1]
    heading_error = angle_difference(compute_heading(prediction), compute_heading(label))

    return {
        "sample_count": int(label.shape[0]),
        "point_count": int(label.shape[0] * label.shape[1]),
        "horizon_steps": int(label.shape[1]),
        "rmse_x": float(np.sqrt(np.mean(diff[:, :, 0] ** 2))),
        "rmse_y": float(np.sqrt(np.mean(diff[:, :, 1] ** 2))),
        "rmse_2d": float(np.sqrt(np.mean(point_error**2))),
        "legacy_rmse_xy_mean": float(np.sqrt(np.mean(diff**2))),
        "mae_x": float(np.mean(np.abs(diff[:, :, 0]))),
        "mae_y": float(np.mean(np.abs(diff[:, :, 1]))),
        "ade": float(np.mean(point_error)),
        "fde": float(np.mean(final_error)),
        "fde_rmse": float(np.sqrt(np.mean(final_error**2))),
        "ahe_rad": float(np.mean(heading_error)),
        "ahe_deg": float(np.degrees(np.mean(heading_error))),
        "fhe_rad": float(np.mean(heading_error[:, -1])),
        "fhe_deg": float(np.degrees(np.mean(heading_error[:, -1]))),
        "bias_x": float(np.mean(diff[:, :, 0])),
        "bias_y": float(np.mean(diff[:, :, 1])),
        "median_error": float(np.percentile(point_error, 50)),
        "p90_error": float(np.percentile(point_error, 90)),
        "p95_error": float(np.percentile(point_error, 95)),
        "p99_error": float(np.percentile(point_error, 99)),
        "max_error": float(np.max(point_error)),
        "rmse_x_by_step": np.round(np.sqrt(np.mean(diff[:, :, 0] ** 2, axis=0)), 6).tolist(),
        "rmse_y_by_step": np.round(np.sqrt(np.mean(diff[:, :, 1] ** 2, axis=0)), 6).tolist(),
        "rmse_2d_by_step": np.round(np.sqrt(np.mean(point_error**2, axis=0)), 6).tolist(),
        "ade_by_step": np.round(np.mean(point_error, axis=0), 6).tolist(),
    }


def relative_path(path):
    try:
        return str(Path(path).resolve().relative_to(BASE_DIR))
    except ValueError:
        return str(path)


def scenario_name_from_dir(path):
    name = Path(path).name
    return name if name.lower().startswith("scenario") else "Scenario0"


def formatted_specs(spec_templates, scenario_name):
    return [
        (
            model_name,
            prediction_template.format(scenario=scenario_name),
            label_template.format(scenario=scenario_name),
        )
        for model_name, prediction_template, label_template in spec_templates
    ]


def add_prediction_variants(specs, llm_dir):
    """Include dated reruns such as *_0503.json as separate model rows."""
    expanded_specs = list(specs)
    seen_prediction_files = {prediction_file for _, prediction_file, _ in expanded_specs}

    for model_name, prediction_file, label_file in specs:
        base_path = llm_dir / prediction_file
        if not base_path.exists():
            continue

        stem = base_path.stem
        for variant_path in sorted(llm_dir.glob(f"{stem}_*.json")):
            if variant_path.name in seen_prediction_files:
                continue
            suffix = variant_path.stem[len(stem) + 1 :]
            expanded_specs.append((f"{model_name}_{suffix}", variant_path.name, label_file))
            seen_prediction_files.add(variant_path.name)

    return expanded_specs


def scenario_specs(scenario_name, llm_dir):
    if scenario_name == "Scenario0":
        return add_prediction_variants(SCENARIO0_MODEL_SPECS, llm_dir)

    candidate_specs = [
        formatted_specs(ALL_OUTPUT_MODEL_SPECS, scenario_name),
        formatted_specs(MODEL_SPECS, scenario_name),
    ]
    for specs in candidate_specs:
        if any(
            (llm_dir / prediction_file).exists() and (llm_dir / label_file).exists()
            for _, prediction_file, label_file in specs
        ):
            return add_prediction_variants(specs, llm_dir)

    return add_prediction_variants(formatted_specs(MODEL_SPECS, scenario_name), llm_dir)


def extract_llm_records(path):
    data = load_json(path)
    records = []
    stats = {
        "total": len(data),
        "parse_failed": 0,
        "invalid_horizon": 0,
    }

    for index, entry in enumerate(data):
        trajectory = parse_trajectory_from_output(entry.get("output"))
        if trajectory is None:
            stats["parse_failed"] += 1
        elif len(trajectory) != EXPECTED_HORIZON_STEPS:
            stats["invalid_horizon"] += 1
            trajectory = None

        records.append(
            {
                "index": index,
                "trajectory": trajectory,
                "input": entry.get("input", ""),
                "output": entry.get("output", ""),
                "load_state": extract_load_state(entry.get("input", "")),
                "nearest_mining_truck_load_state": infer_nearest_mining_truck_load_state(entry.get("output", "")),
                "label_intention": infer_intention_from_text(entry.get("output", "")),
            }
        )

    return records, stats


def load_scenario_metadata(scenario_dir, scenario_name):
    metadata_files = [
        f"MiningCoT_{scenario_name}_test_5000.json",
        f"MiningCoT_{scenario_name}_test_1000.json",
        "MiningCoT_test_5000.json",
        f"LC-LLM_CoT_{scenario_name}_test_5000.json",
        f"LC-LLM_CoT_{scenario_name}_test_1000.json",
        "LC-LLM_CoT_test_5000.json",
    ]
    metadata = []
    for file_name in metadata_files:
        path = scenario_dir / file_name
        if not path.exists():
            continue
        data = load_json(path)
        for index, entry in enumerate(data):
            metadata.append(
                {
                    "index": index,
                    "load_state": extract_load_state(entry.get("input", "")),
                    "label_intention": infer_intention_from_text(entry.get("output", "")),
                }
            )
        if any(item["load_state"] is not None for item in metadata):
            return metadata
    return metadata


def metadata_by_index(metadata):
    return {item["index"]: item for item in metadata}


def input_key(record):
    return normalize_text(record.get("input", ""))


def group_records_by_input(records):
    grouped = {}
    missing_input = 0

    for record in records:
        key = input_key(record)
        if not key:
            missing_input += 1
            continue
        grouped.setdefault(key, []).append(record)

    duplicate_inputs = sum(max(0, len(group) - 1) for group in grouped.values())
    return grouped, {"duplicate_inputs": duplicate_inputs, "missing_inputs": missing_input}


def extract_classical_records(path):
    data = load_json(path)
    records = []
    stats = {
        "total": len(data),
        "label_parse_failed": 0,
        "prediction_parse_failed": 0,
        "invalid_horizon": 0,
    }

    for index, entry in enumerate(data):
        label = normalize_trajectory(entry.get("label") or entry.get("Label"))
        prediction = normalize_trajectory(entry.get("prediction") or entry.get("Prediction"))

        if label is None:
            stats["label_parse_failed"] += 1
        if prediction is None:
            stats["prediction_parse_failed"] += 1
        if label is not None and prediction is not None and len(label) != len(prediction):
            stats["invalid_horizon"] += 1
            min_len = min(len(label), len(prediction))
            label = label[:min_len]
            prediction = prediction[:min_len]

        records.append(
            {
                "index": index,
                "label": label,
                "prediction": prediction,
                "metric_steps": entry.get("metric_steps"),
            }
        )

    return records, stats


def align_llm_records(label_records, prediction_records, distance_threshold, sample_metadata=None):
    stats = {
        "paired": 0,
        "kept": 0,
        "label_count": len(label_records),
        "prediction_count": len(prediction_records),
        "unmatched_labels": 0,
        "unmatched_predictions": 0,
        "label_duplicate_inputs": 0,
        "prediction_duplicate_inputs": 0,
        "label_missing_inputs": 0,
        "prediction_missing_inputs": 0,
        "label_parse_failed": 0,
        "prediction_parse_failed": 0,
        "both_parse_failed": 0,
        "length_mismatch": 0,
        "too_short": 0,
        "prediction_jump_filtered": 0,
    }
    aligned = []
    sample_metadata = sample_metadata or {}

    labels_by_input, label_input_stats = group_records_by_input(label_records)
    predictions_by_input, prediction_input_stats = group_records_by_input(prediction_records)
    stats["label_duplicate_inputs"] = label_input_stats["duplicate_inputs"]
    stats["prediction_duplicate_inputs"] = prediction_input_stats["duplicate_inputs"]
    stats["label_missing_inputs"] = label_input_stats["missing_inputs"]
    stats["prediction_missing_inputs"] = prediction_input_stats["missing_inputs"]

    prediction_offsets = {key: 0 for key in predictions_by_input}
    matched_pairs = []
    unmatched_labels = 0

    for label_record in label_records:
        key = input_key(label_record)
        if not key:
            continue
        predictions = predictions_by_input.get(key, [])
        offset = prediction_offsets.get(key, 0)
        if offset >= len(predictions):
            unmatched_labels += 1
            continue
        matched_pairs.append((label_record, predictions[offset]))
        prediction_offsets[key] = offset + 1

    unmatched_predictions = sum(
        len(predictions) - prediction_offsets.get(key, 0) for key, predictions in predictions_by_input.items()
    )
    stats["paired"] = len(matched_pairs)
    stats["unmatched_labels"] = unmatched_labels
    stats["unmatched_predictions"] = unmatched_predictions

    for label_record, prediction_record in matched_pairs:
        label = label_record["trajectory"]
        prediction = prediction_record["trajectory"]

        if label is None and prediction is None:
            stats["both_parse_failed"] += 1
            continue
        if label is None:
            stats["label_parse_failed"] += 1
            continue
        if prediction is None:
            stats["prediction_parse_failed"] += 1
            continue
        if len(label) != len(prediction):
            stats["length_mismatch"] += 1
            continue
        if len(label) < 2:
            stats["too_short"] += 1
            continue

        jumps = np.linalg.norm(np.diff(np.asarray(prediction, dtype=float), axis=0), axis=1)
        if np.any(jumps > distance_threshold):
            stats["prediction_jump_filtered"] += 1
            continue

        aligned.append(
            {
                "index": label_record["index"],
                "label": label,
                "prediction": prediction,
                "load_state": label_record.get("load_state") or sample_metadata.get(label_record["index"], {}).get("load_state"),
                "nearest_mining_truck_load_state": label_record.get("nearest_mining_truck_load_state"),
                "prediction_nearest_mining_truck_load_state": prediction_record.get("nearest_mining_truck_load_state"),
                "prediction_load_state": prediction_record.get("load_state"),
                "label_intention": intention_name(label_record.get("label_intention"))
                or sample_metadata.get(label_record["index"], {}).get("label_intention"),
                "prediction_intention": infer_intention_from_text(prediction_record["output"]),
                "label_input": label_record.get("input", ""),
                "prediction_input": prediction_record.get("input", ""),
            }
        )

    stats["kept"] = len(aligned)
    return aligned, stats


def align_classical_records(records, sample_metadata=None):
    stats = {
        "paired": len(records),
        "kept": 0,
        "label_parse_failed": 0,
        "prediction_parse_failed": 0,
        "both_parse_failed": 0,
        "too_short": 0,
    }
    aligned = []
    sample_metadata = sample_metadata or {}

    for record in records:
        label = record["label"]
        prediction = record["prediction"]

        if label is None and prediction is None:
            stats["both_parse_failed"] += 1
            continue
        if label is None:
            stats["label_parse_failed"] += 1
            continue
        if prediction is None:
            stats["prediction_parse_failed"] += 1
            continue
        if len(label) < 2:
            stats["too_short"] += 1
            continue

        metadata = sample_metadata.get(record["index"], {})
        aligned.append(
            {
                "index": record["index"],
                "label": label,
                "prediction": prediction,
                "load_state": metadata.get("load_state"),
                "prediction_load_state": None,
            }
        )

    stats["kept"] = len(aligned)
    return aligned, stats


def normalize_text(text):
    return re.sub(r"\s+", " ", str(text or "")).strip().lower()


def extract_load_state(text):
    loaded_values = re.findall(r"--\s*Loaded\s*:\s*(Not\s+Full\s+Loaded|Full\s+Loaded)", str(text or ""), re.IGNORECASE)
    if any(normalize_text(value) == "full loaded" for value in loaded_values):
        return "full_loaded"
    if loaded_values:
        return "not_full_loaded"
    return None


def infer_nearest_mining_truck_load_state(text):
    normalized = normalize_text(text)
    if NEAREST_MINING_TRUCK_FULLY_LOADED_PHRASE in normalized:
        return "full_loaded"
    return "not_full_loaded"


def infer_intention_from_text(text):
    normalized = normalize_text(text)
    if re.search(r"\bturn\s+left\b|\bleftward\b|\bleft[- ]bending\b|\bleft\s+turn\b", normalized):
        return "left"
    if re.search(r"\bturn\s+right\b|\brightward\b|\brightward\s+curvature\b|\bright\s+turn\b", normalized):
        return "right"
    if re.search(r"\bgo\s+straight\b|\bstable\s+lane[- ]keeping\b|\bcurrent\s+lane\b|\breference\s+path\b", normalized):
        return "straight"
    return None


def intention_name(intention):
    if intention is None:
        return None
    if isinstance(intention, str):
        normalized = normalize_text(intention)
        return normalized if normalized in INTENTION_CLASSES else None
    return None


def safe_divide(numerator, denominator):
    return numerator / denominator if denominator else None


def compute_intention_prediction_metrics(aligned):
    confusion = {label: {pred: 0 for pred in INTENTION_CLASSES} for label in INTENTION_CLASSES}
    support = {label: 0 for label in INTENTION_CLASSES}
    predicted_support = {label: 0 for label in INTENTION_CLASSES}
    valid = 0
    label_parse_failed = 0
    prediction_parse_failed = 0

    for sample in aligned:
        label = intention_name(sample.get("label_intention"))
        prediction = intention_name(sample.get("prediction_intention"))

        if label is None:
            label_parse_failed += 1
        if prediction is None:
            prediction_parse_failed += 1
        if label is None or prediction is None:
            continue

        valid += 1
        support[label] += 1
        predicted_support[prediction] += 1
        confusion[label][prediction] += 1

    per_class = {}
    for label in INTENTION_CLASSES:
        true_positive = confusion[label][label]
        false_positive = sum(confusion[actual][label] for actual in INTENTION_CLASSES if actual != label)
        false_negative = sum(confusion[label][predicted] for predicted in INTENTION_CLASSES if predicted != label)
        precision = safe_divide(true_positive, true_positive + false_positive)
        recall = safe_divide(true_positive, true_positive + false_negative)
        if precision is None or recall is None:
            f1 = None
        elif precision + recall == 0:
            f1 = 0.0
        else:
            f1 = 2 * precision * recall / (precision + recall)
        per_class[label] = {
            "precision": precision,
            "recall": recall,
            "f1": f1,
            "support": support[label],
            "predicted_support": predicted_support[label],
        }

    macro_avg = {}
    for metric_name in ("precision", "recall", "f1"):
        values = [per_class[label][metric_name] for label in INTENTION_CLASSES]
        macro_avg[metric_name] = float(np.mean([value for value in values if value is not None])) if any(
            value is not None for value in values
        ) else None

    return {
        "valid_count": valid,
        "label_parse_failed": label_parse_failed,
        "prediction_parse_failed": prediction_parse_failed,
        "classes": INTENTION_CLASSES,
        "per_class": per_class,
        "macro_avg": macro_avg,
        "confusion_matrix": confusion,
    }


def compute_binary_classification_metrics(aligned, label_key, prediction_key, classes):
    confusion = {label: {pred: 0 for pred in classes} for label in classes}
    valid = 0
    label_parse_failed = 0
    prediction_parse_failed = 0

    for sample in aligned:
        label = sample.get(label_key)
        prediction = sample.get(prediction_key)

        if label not in classes:
            label_parse_failed += 1
        if prediction not in classes:
            prediction_parse_failed += 1
        if label not in classes or prediction not in classes:
            continue

        valid += 1
        confusion[label][prediction] += 1

    per_class = {}
    for label in classes:
        true_positive = confusion[label][label]
        false_positive = sum(confusion[actual][label] for actual in classes if actual != label)
        false_negative = sum(confusion[label][predicted] for predicted in classes if predicted != label)
        precision = safe_divide(true_positive, true_positive + false_positive)
        recall = safe_divide(true_positive, true_positive + false_negative)
        if precision is None or recall is None:
            f1 = None
        elif precision + recall == 0:
            f1 = 0.0
        else:
            f1 = 2 * precision * recall / (precision + recall)
        per_class[label] = {
            "precision": precision,
            "recall": recall,
            "f1": f1,
        }

    macro_avg = {}
    for metric_name in ("precision", "recall", "f1"):
        values = [per_class[label][metric_name] for label in classes]
        macro_avg[metric_name] = float(np.mean([value for value in values if value is not None])) if any(
            value is not None for value in values
        ) else None

    match_count = sum(confusion[label][label] for label in classes)
    return {
        "valid_count": valid,
        "match_count": match_count,
        "accuracy": safe_divide(match_count, valid),
        "label_parse_failed": label_parse_failed,
        "prediction_parse_failed": prediction_parse_failed,
        "per_class": per_class,
        "macro_avg": macro_avg,
        "confusion_matrix": confusion,
    }


def compute_load_state_prediction_metrics(aligned):
    return compute_binary_classification_metrics(
        aligned,
        "nearest_mining_truck_load_state",
        "prediction_nearest_mining_truck_load_state",
        LOAD_STATES,
    )


def compute_point_bias_metrics(aligned):
    label = np.asarray([record["label"] for record in aligned], dtype=float)
    prediction = np.asarray([record["prediction"] for record in aligned], dtype=float)
    diff = prediction - label
    point_error = np.linalg.norm(diff, axis=2)

    metrics = {}
    for point_index in range(diff.shape[1]):
        step = point_index + 1
        dx = diff[:, point_index, 0]
        dy = diff[:, point_index, 1]
        distance = point_error[:, point_index]
        metrics[f"t{step}"] = {
            "mae_x": float(np.mean(np.abs(dx))),
            "mae_y": float(np.mean(np.abs(dy))),
            "bias_x": float(np.mean(dx)),
            "bias_y": float(np.mean(dy)),
            "rmse_x": float(np.sqrt(np.mean(dx**2))),
            "rmse_y": float(np.sqrt(np.mean(dy**2))),
            "mean_2d_error": float(np.mean(distance)),
            "rmse_2d": float(np.sqrt(np.mean(distance**2))),
            "p95_2d_error": float(np.percentile(distance, 95)),
        }
    return metrics


def evaluate_sample_mean_trajectory_metrics(aligned):
    if not aligned:
        raise ValueError("No valid aligned trajectories are available.")

    sample_metrics = []
    all_point_errors = []
    for record in aligned:
        label = np.asarray(record["label"], dtype=float)
        prediction = np.asarray(record["prediction"], dtype=float)
        diff = prediction - label
        point_error = np.linalg.norm(diff, axis=1)
        all_point_errors.extend(point_error.tolist())

        if len(label) >= 2:
            heading_error = angle_difference(compute_heading(prediction[None, :, :]), compute_heading(label[None, :, :]))[0]
            ahe_deg = float(np.degrees(np.mean(heading_error)))
            fhe_deg = float(np.degrees(heading_error[-1]))
        else:
            ahe_deg = None
            fhe_deg = None

        sample_metrics.append(
            {
                "rmse_x": float(np.sqrt(np.mean(diff[:, 0] ** 2))),
                "rmse_y": float(np.sqrt(np.mean(diff[:, 1] ** 2))),
                "rmse_2d": float(np.sqrt(np.mean(point_error**2))),
                "mae_x": float(np.mean(np.abs(diff[:, 0]))),
                "mae_y": float(np.mean(np.abs(diff[:, 1]))),
                "ade": float(np.mean(point_error)),
                "fde": float(point_error[-1]),
                "fde_rmse": float(point_error[-1]),
                "ahe_deg": ahe_deg,
                "fhe_deg": fhe_deg,
                "bias_x": float(np.mean(diff[:, 0])),
                "bias_y": float(np.mean(diff[:, 1])),
            }
        )

    def mean_metric(metric_name):
        values = [item[metric_name] for item in sample_metrics if item[metric_name] is not None]
        return float(np.mean(values)) if values else None

    point_errors = np.asarray(all_point_errors, dtype=float)
    first_label = aligned[0]["label"]
    return {
        "sample_count": len(aligned),
        "point_count": int(len(aligned) * len(first_label)),
        "horizon_steps": int(len(first_label)),
        "aggregation": "sample_mean",
        "rmse_x": mean_metric("rmse_x"),
        "rmse_y": mean_metric("rmse_y"),
        "rmse_2d": mean_metric("rmse_2d"),
        "mae_x": mean_metric("mae_x"),
        "mae_y": mean_metric("mae_y"),
        "ade": mean_metric("ade"),
        "fde": mean_metric("fde"),
        "fde_rmse": mean_metric("fde_rmse"),
        "ahe_deg": mean_metric("ahe_deg"),
        "fhe_deg": mean_metric("fhe_deg"),
        "bias_x": mean_metric("bias_x"),
        "bias_y": mean_metric("bias_y"),
        "median_error": float(np.percentile(point_errors, 50)),
        "p90_error": float(np.percentile(point_errors, 90)),
        "p95_error": float(np.percentile(point_errors, 95)),
        "p99_error": float(np.percentile(point_errors, 99)),
        "max_error": float(np.max(point_errors)),
    }


def evaluate_load_state_subsets(aligned, include_intention=False):
    results = {}
    for load_state in LOAD_STATES:
        subset = [record for record in aligned if record.get("load_state") == load_state]
        if subset:
            metrics = evaluate_sample_mean_trajectory_metrics(subset)
            metrics["point_bias_metrics"] = compute_point_bias_metrics(subset)
            if include_intention:
                metrics["intention_prediction"] = compute_intention_prediction_metrics(subset)
        else:
            metrics = None
        results[load_state] = {
            "sample_count": len(subset),
            "metrics": metrics,
        }

    unknown_count = sum(1 for record in aligned if record.get("load_state") not in LOAD_STATES)
    results["unknown"] = {"sample_count": unknown_count, "metrics": None}
    return results


def evaluate_point_deviation(aligned):
    if not aligned:
        raise ValueError("No valid aligned trajectories are available.")

    label = np.asarray([record["label"] for record in aligned], dtype=float)
    prediction = np.asarray([record["prediction"] for record in aligned], dtype=float)
    diff = prediction - label
    point_error = np.linalg.norm(diff, axis=2)
    final_error = point_error[:, -1]
    heading_error = angle_difference(compute_heading(prediction), compute_heading(label))

    return {
        "sample_count": int(label.shape[0]),
        "point_count": int(label.shape[0] * label.shape[1]),
        "horizon_steps": int(label.shape[1]),
        "rmse_x": float(np.sqrt(np.mean(diff[:, :, 0] ** 2))),
        "rmse_y": float(np.sqrt(np.mean(diff[:, :, 1] ** 2))),
        "rmse_2d": float(np.sqrt(np.mean(point_error**2))),
        "legacy_rmse_xy_mean": float(np.sqrt(np.mean(diff**2))),
        "mae_x": float(np.mean(np.abs(diff[:, :, 0]))),
        "mae_y": float(np.mean(np.abs(diff[:, :, 1]))),
        "ade": float(np.mean(point_error)),
        "fde": float(np.mean(final_error)),
        "fde_rmse": float(np.sqrt(np.mean(final_error**2))),
        "ahe_rad": float(np.mean(heading_error)),
        "ahe_deg": float(np.degrees(np.mean(heading_error))),
        "fhe_rad": float(np.mean(heading_error[:, -1])),
        "fhe_deg": float(np.degrees(np.mean(heading_error[:, -1]))),
        "bias_x": float(np.mean(diff[:, :, 0])),
        "bias_y": float(np.mean(diff[:, :, 1])),
        "median_error": float(np.percentile(point_error, 50)),
        "p90_error": float(np.percentile(point_error, 90)),
        "p95_error": float(np.percentile(point_error, 95)),
        "p99_error": float(np.percentile(point_error, 99)),
        "max_error": float(np.max(point_error)),
        "rmse_x_by_step": np.round(np.sqrt(np.mean(diff[:, :, 0] ** 2, axis=0)), 6).tolist(),
        "rmse_y_by_step": np.round(np.sqrt(np.mean(diff[:, :, 1] ** 2, axis=0)), 6).tolist(),
        "rmse_2d_by_step": np.round(np.sqrt(np.mean(point_error**2, axis=0)), 6).tolist(),
        "ade_by_step": np.round(np.mean(point_error, axis=0), 6).tolist(),
    }


def add_common_metrics(result, aligned, include_intention=False):
    result["metrics"]["point_bias_metrics"] = compute_point_bias_metrics(aligned)
    if include_intention:
        result["metrics"]["intention_prediction"] = compute_intention_prediction_metrics(aligned)
    result["metrics"]["load_state_evaluation"] = compute_load_state_prediction_metrics(aligned)


def evaluate_llm_dir(llm_dir, distance_threshold, scenario_name=None):
    scenario_name = scenario_name or scenario_name_from_dir(llm_dir)
    sample_metadata = metadata_by_index(load_scenario_metadata(llm_dir, scenario_name))
    results = {}

    for model_name, prediction_file, label_file in scenario_specs(scenario_name, llm_dir):
        prediction_path = llm_dir / prediction_file
        label_path = llm_dir / label_file
        if not prediction_path.exists() or not label_path.exists():
            continue

        labels, label_stats = extract_llm_records(label_path)
        predictions, prediction_stats = extract_llm_records(prediction_path)
        aligned, alignment = align_llm_records(labels, predictions, distance_threshold, sample_metadata)
        metrics = evaluate_aligned(aligned)

        result = {
            "group": "LLM",
            "scenario": scenario_name,
            "prediction_file": relative_path(prediction_path),
            "label_file": relative_path(label_path),
            "metrics": metrics,
            "alignment": alignment,
            "label_source": label_stats,
            "prediction_source": prediction_stats,
        }
        add_common_metrics(result, aligned, include_intention=True)
        results[model_name] = result

    return results


def evaluate_classical_dir(classical_dir, scenario_name):
    sample_metadata = metadata_by_index(load_scenario_metadata(classical_dir, scenario_name))
    results = {}
    for path in sorted(classical_dir.glob("*_predictions.json")):
        model_name = path.name.replace("_predictions.json", "")
        records, source_stats = extract_classical_records(path)
        aligned, alignment = align_classical_records(records, sample_metadata)
        metrics = evaluate_point_deviation(aligned)

        result = {
            "group": "Classical",
            "scenario": scenario_name,
            "prediction_file": relative_path(path),
            "label_file": relative_path(path),
            "metrics": metrics,
            "alignment": alignment,
            "source": source_stats,
        }
        add_common_metrics(result, aligned, include_intention=False)
        results[model_name] = result

    return results


def percent_delta(value, baseline):
    if baseline in (None, 0):
        return None
    return (value - baseline) / baseline * 100.0


def rank_rows(rows):
    ranked = sorted(rows, key=lambda row: (row["rmse_2d"], row["fde"], row["ade"]))
    for index, row in enumerate(ranked, start=1):
        row["rank"] = index

    by_group = {}
    for row in ranked:
        by_group.setdefault(row["group"], []).append(row)

    best_overall = ranked[0] if ranked else None
    best_classical = by_group.get("Classical", [None])[0]
    best_llm = by_group.get("LLM", [None])[0]

    for row in ranked:
        row["delta_vs_best_rmse2d_pct"] = (
            percent_delta(row["rmse_2d"], best_overall["rmse_2d"]) if best_overall else None
        )
        row["delta_vs_best_classical_rmse2d_pct"] = (
            percent_delta(row["rmse_2d"], best_classical["rmse_2d"]) if best_classical else None
        )
        row["delta_vs_best_llm_rmse2d_pct"] = percent_delta(row["rmse_2d"], best_llm["rmse_2d"]) if best_llm else None

    return ranked


def flatten_result(model_name, result):
    metrics = result["metrics"]
    row = {
        "model": model_name,
        "group": result["group"],
        "scenario": result["scenario"],
        "sample_count": metrics["sample_count"],
        "point_count": metrics["point_count"],
        "horizon_steps": metrics["horizon_steps"],
        "rmse_x": metrics["rmse_x"],
        "rmse_y": metrics["rmse_y"],
        "rmse_2d": metrics["rmse_2d"],
        "mae_x": metrics["mae_x"],
        "mae_y": metrics["mae_y"],
        "ade": metrics["ade"],
        "fde": metrics["fde"],
        "fde_rmse": metrics["fde_rmse"],
        "ahe_deg": metrics["ahe_deg"],
        "fhe_deg": metrics["fhe_deg"],
        "bias_x": metrics["bias_x"],
        "bias_y": metrics["bias_y"],
        "median_error": metrics["median_error"],
        "p90_error": metrics["p90_error"],
        "p95_error": metrics["p95_error"],
        "p99_error": metrics["p99_error"],
        "max_error": metrics["max_error"],
        "kept": result["alignment"]["kept"],
        "paired": result["alignment"]["paired"],
        "label_parse_failed": result["alignment"]["label_parse_failed"],
        "prediction_parse_failed": result["alignment"]["prediction_parse_failed"],
        "prediction_file": result["prediction_file"],
        "label_file": result["label_file"],
    }

    for index, value in enumerate(metrics["rmse_2d_by_step"], start=1):
        row[f"rmse_2d_t{index}"] = value
    for index, value in enumerate(metrics["ade_by_step"], start=1):
        row[f"ade_t{index}"] = value
    for step, values in metrics["point_bias_metrics"].items():
        for metric_name, value in values.items():
            row[f"{step}_{metric_name}"] = value

    intention = metrics.get("intention_prediction")
    if intention is not None:
        row["intention_valid_count"] = intention["valid_count"]
        row["intention_macro_precision"] = intention["macro_avg"]["precision"]
        row["intention_macro_recall"] = intention["macro_avg"]["recall"]
        row["intention_macro_f1"] = intention["macro_avg"]["f1"]
        for label in INTENTION_CLASSES:
            class_metrics = intention["per_class"][label]
            row[f"intention_{label}_precision"] = class_metrics["precision"]
            row[f"intention_{label}_recall"] = class_metrics["recall"]
            row[f"intention_{label}_f1"] = class_metrics["f1"]
            row[f"intention_{label}_support"] = class_metrics["support"]

    load_state = metrics.get("load_state_evaluation")
    if load_state is not None:
        for state in LOAD_STATES:
            state_result = load_state[state]
            row[f"{state}_sample_count"] = state_result["sample_count"]
            state_metrics = state_result["metrics"]
            if state_metrics is None:
                continue
            for metric_name in ("rmse_2d", "ade", "fde", "fde_rmse", "p95_error", "ahe_deg", "fhe_deg"):
                row[f"{state}_{metric_name}"] = state_metrics.get(metric_name)
            state_intention = state_metrics.get("intention_prediction")
            if state_intention is not None:
                row[f"{state}_intention_macro_precision"] = state_intention["macro_avg"]["precision"]
                row[f"{state}_intention_macro_recall"] = state_intention["macro_avg"]["recall"]
                row[f"{state}_intention_macro_f1"] = state_intention["macro_avg"]["f1"]
        row["unknown_load_state_sample_count"] = load_state["unknown"]["sample_count"]

    return row


def flatten_ampilot_metrix_result(model_name, result):
    metrics = result["metrics"]
    is_mining_cot_model = re.match(r"^MiningCoT(?:7B|13B)(?:_|$)", model_name) is not None
    is_llm_planner = "Mining" in model_name or "LC-LLM" in model_name
    row = {
        "scenario": result["scenario"],
        "group": result["group"],
        "model": model_name,
        "sample_count": metrics["sample_count"],
        "RMSE": metrics.get("rmse_2d"),
        "ADE": metrics.get("ade"),
        "AHE": metrics.get("ahe_rad"),
        "FDE": metrics.get("fde"),
        "FHE": metrics.get("fhe_rad"),
        "load_state_evaluation": format_load_state_evaluation(metrics) if is_mining_cot_model else "",
        "intention_prediction": format_intention_prediction(metrics) if is_llm_planner else "",
    }

    for index in range(EXPECTED_HORIZON_STEPS):
        row[f"rmse_x_t{index + 1}"] = metrics["rmse_x_by_step"][index] if index < len(metrics["rmse_x_by_step"]) else None
    for index in range(EXPECTED_HORIZON_STEPS):
        row[f"rmse_y_t{index + 1}"] = metrics["rmse_y_by_step"][index] if index < len(metrics["rmse_y_by_step"]) else None

    return row


def fmt_metric_text(value):
    return "NA" if value is None else f"{float(value):.4f}"


def format_load_state_evaluation(metrics):
    load_state = metrics.get("load_state_evaluation") or {}
    if not load_state or not load_state.get("valid_count"):
        return "Not available"

    per_class = load_state["per_class"]
    full_loaded = per_class["full_loaded"]
    not_full_loaded = per_class["not_full_loaded"]
    return "\n".join(
        [
            f"Accuracy: {fmt_metric_text(load_state.get('accuracy'))}",
            "M_FL:",
            f"  Precision: {fmt_metric_text(full_loaded.get('precision'))}",
            f"  Recall: {fmt_metric_text(full_loaded.get('recall'))}",
            f"  F1 Score: {fmt_metric_text(full_loaded.get('f1'))}",
            "M_NFL:",
            f"  Precision: {fmt_metric_text(not_full_loaded.get('precision'))}",
            f"  Recall: {fmt_metric_text(not_full_loaded.get('recall'))}",
            f"  F1 Score: {fmt_metric_text(not_full_loaded.get('f1'))}",
        ]
    )


def format_intention_prediction(metrics):
    intention = metrics.get("intention_prediction")
    if intention is None:
        return "Not available"

    per_class = intention["per_class"]

    def class_values(class_name):
        class_metrics = per_class[class_name]
        return [
            f"{class_name}:",
            f"  Precision: {fmt_metric_text(class_metrics['precision'])}",
            f"  Recall: {fmt_metric_text(class_metrics['recall'])}",
            f"  F1 Score: {fmt_metric_text(class_metrics['f1'])}",
        ]

    macro_avg = intention["macro_avg"]
    lines = []
    for class_name in ("straight", "left", "right"):
        lines.extend(class_values(class_name))
    lines.extend(
        [
            "all:",
            f"  Precision: {fmt_metric_text(macro_avg['precision'])}",
            f"  Recall: {fmt_metric_text(macro_avg['recall'])}",
            f"  F1 Score: {fmt_metric_text(macro_avg['f1'])}",
            f"  Macro Avg: P={fmt_metric_text(macro_avg['precision'])}, R={fmt_metric_text(macro_avg['recall'])}, F1={fmt_metric_text(macro_avg['f1'])}",
        ]
    )
    return "\n".join(lines)


def fmt_float(value, digits=4):
    if value is None or value == "":
        return "-"
    return f"{float(value):.{digits}f}"


def fmt_pct(value):
    if value is None or value == "":
        return "-"
    return f"{float(value):+.2f}%"


def print_comparison(rows):
    rows = rank_rows(rows)
    if not rows:
        print("No valid results found.")
        return

    scenario = rows[0]["scenario"]
    print(f"\n{scenario} prediction comparison, sorted by RMSE_2D (lower is better)")
    header = (
        f"{'Rank':>4}  {'Group':<9}  {'Model':<18}  {'N':>5}  "
        f"{'RMSE_2D':>8}  {'ADE':>8}  {'FDE':>8}  {'P95':>8}  {'AHE_deg':>8}  {'vs Best':>9}"
    )
    print(header)
    print("-" * len(header))
    for row in rows:
        print(
            f"{row['rank']:>4}  {row['group']:<9}  {row['model']:<18}  {row['sample_count']:>5}  "
            f"{fmt_float(row['rmse_2d']):>8}  {fmt_float(row['ade']):>8}  {fmt_float(row['fde']):>8}  "
            f"{fmt_float(row['p95_error']):>8}  {fmt_float(row['ahe_deg'], 2):>8}  "
            f"{fmt_pct(row['delta_vs_best_rmse2d_pct']):>9}"
        )

    print("\nBest by group")
    for group in ("Classical", "LLM"):
        group_rows = [row for row in rows if row["group"] == group]
        if not group_rows:
            continue
        best = group_rows[0]
        print(
            f"- {group}: {best['model']} | RMSE_2D={fmt_float(best['rmse_2d'])}, "
            f"ADE={fmt_float(best['ade'])}, FDE={fmt_float(best['fde'])}"
        )

    print("\nPer-step RMSE_2D")
    step_header = f"{'Model':<18}  {'Group':<9}  {'t1':>8}  {'t2':>8}  {'t3':>8}  {'t4':>8}"
    print(step_header)
    print("-" * len(step_header))
    for row in rows:
        print(
            f"{row['model']:<18}  {row['group']:<9}  "
            f"{fmt_float(row.get('rmse_2d_t1')):>8}  {fmt_float(row.get('rmse_2d_t2')):>8}  "
            f"{fmt_float(row.get('rmse_2d_t3')):>8}  {fmt_float(row.get('rmse_2d_t4')):>8}"
        )

    print("\nIntention prediction")
    intention_header = (
        f"{'Model':<18}  {'Group':<9}  {'Valid':>5}  {'Macro-P':>8}  {'Macro-R':>8}  {'Macro-F1':>8}"
    )
    print(intention_header)
    print("-" * len(intention_header))
    for row in [item for item in rows if item["group"] == "LLM"]:
        print(
            f"{row['model']:<18}  {row['group']:<9}  {row.get('intention_valid_count', 0):>5}  "
            f"{fmt_float(row.get('intention_macro_precision')):>8}  "
            f"{fmt_float(row.get('intention_macro_recall')):>8}  "
            f"{fmt_float(row.get('intention_macro_f1')):>8}"
        )

    print("\nLoad-state RMSE_2D")
    load_header = f"{'Model':<18}  {'Group':<9}  {'FL_N':>5}  {'FL':>8}  {'NFL_N':>5}  {'NFL':>8}"
    print(load_header)
    print("-" * len(load_header))
    for row in rows:
        print(
            f"{row['model']:<18}  {row['group']:<9}  "
            f"{row.get('full_loaded_sample_count', 0):>5}  {fmt_float(row.get('full_loaded_rmse_2d')):>8}  "
            f"{row.get('not_full_loaded_sample_count', 0):>5}  {fmt_float(row.get('not_full_loaded_rmse_2d')):>8}"
        )


def print_ampilot_metrix(rows):
    if not rows:
        print("No valid results found.")
        return

    scenario = rows[0]["scenario"]
    print(f"\n{scenario} AMPilot-Metrix")
    header = (
        f"{'Group':<9}  {'Model':<18}  {'N':>5}  "
        f"{'RMSE':>8}  {'ADE':>8}  {'AHE':>8}  {'FDE':>8}  {'FHE':>8}  "
        f"{'X_t1':>8}  {'X_t2':>8}  {'X_t3':>8}  {'X_t4':>8}  "
        f"{'Y_t1':>8}  {'Y_t2':>8}  {'Y_t3':>8}  {'Y_t4':>8}"
    )
    print(header)
    print("-" * len(header))
    for row in rows:
        print(
            f"{row['group']:<9}  {row['model']:<18}  {row['sample_count']:>5}  "
            f"{fmt_float(row.get('RMSE')):>8}  "
            f"{fmt_float(row.get('ADE')):>8}  {fmt_float(row.get('AHE')):>8}  "
            f"{fmt_float(row.get('FDE')):>8}  {fmt_float(row.get('FHE')):>8}  "
            f"{fmt_float(row.get('rmse_x_t1')):>8}  {fmt_float(row.get('rmse_x_t2')):>8}  "
            f"{fmt_float(row.get('rmse_x_t3')):>8}  {fmt_float(row.get('rmse_x_t4')):>8}  "
            f"{fmt_float(row.get('rmse_y_t1')):>8}  {fmt_float(row.get('rmse_y_t2')):>8}  "
            f"{fmt_float(row.get('rmse_y_t3')):>8}  {fmt_float(row.get('rmse_y_t4')):>8}"
        )


def write_csv(rows, path):
    fieldnames = sorted({key for row in rows for key in row})
    preferred = AMPILOT_METRIX_COLUMNS
    ordered = preferred + [field for field in fieldnames if field not in preferred]
    path.parent.mkdir(parents=True, exist_ok=True)
    output_path = path
    for attempt in range(2, 100):
        try:
            with open(output_path, "w", encoding="utf-8", newline="") as file:
                writer = csv.DictWriter(file, fieldnames=ordered)
                writer.writeheader()
                writer.writerows(rows)
            return output_path
        except PermissionError:
            output_path = path.with_name(f"{path.stem}_{attempt}{path.suffix}")

    raise PermissionError(f"Could not write CSV to {path} or fallback paths.")


def write_xlsx(rows, path):
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "AMPilot-Metrix"

    headers = AMPILOT_METRIX_COLUMNS
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wrap_columns = {"load_state_evaluation", "intention_prediction"}
    for column_index, header in enumerate(headers, start=1):
        column_letter = get_column_letter(column_index)
        if header in wrap_columns:
            worksheet.column_dimensions[column_letter].width = 42 if header == "load_state_evaluation" else 48
        elif header in {"model", "scenario"}:
            worksheet.column_dimensions[column_letter].width = 18
        else:
            worksheet.column_dimensions[column_letter].width = 12

        for cell in worksheet[column_letter][1:]:
            cell.alignment = Alignment(vertical="top", wrap_text=header in wrap_columns)

    output_path = path
    for attempt in range(2, 100):
        try:
            workbook.save(output_path)
            return output_path
        except PermissionError:
            output_path = path.with_name(f"{path.stem}_{attempt}{path.suffix}")

    raise PermissionError(f"Could not write XLSX to {path} or fallback paths.")


def existing_directory(path):
    path = Path(path)
    if not path.is_dir():
        raise argparse.ArgumentTypeError(f"Directory does not exist: {path}")
    return path


def parse_args():
    parser = argparse.ArgumentParser(
        description="Evaluate AMPilot scenario LLM and classical trajectory predictions.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
        epilog=(
            "Example: python evaluate_predict0415_scenarios.py "
            "--llm-dir ../AMPilot-Dataset/TestDataset/Scenario2 "
            "--classical-dir ../AMPilot-Dataset/TestDataset/Scenario2 "
            "--output-dir ../AMPilot-Dataset/TestDataset/Scenario2Comparison"
        ),
    )
    parser.add_argument("--llm-dir", type=existing_directory, help="Directory containing LLM prediction/test JSON files.")
    parser.add_argument(
        "--classical-dir",
        type=existing_directory,
        help="Classical prediction directory containing *_predictions.json files.",
    )
    parser.add_argument("--output-dir", type=Path, required=True, help="Directory where JSON and XLSX metric files are written.")
    parser.add_argument(
        "--scenario",
        help="Scenario name used in output rows and filenames. Defaults to the input directory name, such as Scenario2.",
    )
    parser.add_argument("--llm-distance-threshold", type=float, default=LLM_DISTANCE_THRESHOLD)
    parser.add_argument("--skip-llm", action="store_true", help="Do not evaluate LLM prediction files.")
    parser.add_argument("--skip-classical", action="store_true", help="Do not evaluate classical *_predictions.json files.")
    args = parser.parse_args()

    if args.skip_llm and args.skip_classical:
        parser.error("At least one evaluator must run; remove --skip-llm or --skip-classical.")
    if not args.skip_llm and args.llm_dir is None:
        parser.error("--llm-dir is required unless --skip-llm is set.")
    if not args.skip_classical and args.classical_dir is None:
        parser.error("--classical-dir is required unless --skip-classical is set.")
    if args.scenario is None and args.llm_dir is None and args.classical_dir is None:
        parser.error("--scenario is required when no input directory is available for scenario inference.")

    return args


def main():
    args = parse_args()
    scenario_source_dir = args.llm_dir or args.classical_dir
    scenario_name = args.scenario or scenario_name_from_dir(scenario_source_dir)
    results = {}

    if not args.skip_llm:
        results.update(evaluate_llm_dir(args.llm_dir, args.llm_distance_threshold, scenario_name))
    if not args.skip_classical:
        results.update(evaluate_classical_dir(args.classical_dir, scenario_name))

    rows = [flatten_ampilot_metrix_result(model_name, result) for model_name, result in results.items()]
    rows = sorted(rows, key=lambda row: (row["RMSE"], row["FDE"], row["ADE"]))

    args.output_dir.mkdir(parents=True, exist_ok=True)
    json_out = args.output_dir / f"{scenario_name.lower()}_ampilot_metrix.json"
    xlsx_out = args.output_dir / f"{scenario_name.lower()}_ampilot_metrix.xlsx"

    with open(json_out, "w", encoding="utf-8") as file:
        json.dump(rows, file, indent=2, ensure_ascii=False)
    xlsx_out = write_xlsx(rows, xlsx_out)

    print_ampilot_metrix(rows)
    print(f"\nSaved JSON: {json_out}")
    print(f"Saved XLSX: {xlsx_out}")


if __name__ == "__main__":
    main()
